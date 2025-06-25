import time
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# --------------------------- 配置区 ---------------------------
URL = "https://openaccess.thecvf.com/CVPR2025?day=all"
OUTPUT_FILE = "cvpr2025_papers.xlsx"
HEADERS      = {"User-Agent": "Mozilla/5.0"}   # 简易反爬头
SLEEP_BETWEEN_REQUESTS = 1.0                   # 请求间隔（秒）
# --------------------------------------------------------------

def fetch_html(url: str) -> str:
    """获取网页 HTML 文本"""
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    return resp.text

def parse_papers(html: str):
    """解析所有论文的标题与作者"""
    soup     = BeautifulSoup(html, "lxml")
    for dt in soup.select("dt.ptitle"):
        title = dt.get_text(strip=True)
        dd    = dt.find_next_sibling("dd")
        # 只提取 <a> 内文本，去掉 html 里额外的逗号/换行
        author_list = [a.get_text(strip=True) for a in dd.select("a")]
        authors     = ", ".join(author_list)
        yield title, authors

def build_excel(rows):
    """根据 rows 构建带格式的 Excel"""
    wb   = Workbook()
    ws   = wb.active
    font = Font(name="Sitka Text", size=11)
    align= Alignment(wrap_text=True, horizontal="center", vertical="center")

    # 表头
    headers = ["Conference/Journal", "Title", "Authors"]
    ws.append(headers)
    
    # 数据行
    for title, authors in rows:
        ws.append(["CVPR2025", title, authors])

    # 列宽
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 140
    ws.column_dimensions["C"].width = 140

    # 统一字体
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=3):
        for cell in row:
            cell.font = font
            cell.alignment = align

    wb.save(OUTPUT_FILE)
    print(f"Saved → {OUTPUT_FILE}")

def main():
    print("Fetching page …")
    html = fetch_html(URL)
    time.sleep(SLEEP_BETWEEN_REQUESTS)        # 友好延时

    print("Parsing papers …")
    papers = list(parse_papers(html))
    print(f"Found {len(papers)} papers.")

    print("Building Excel …")
    build_excel(papers)

if __name__ == "__main__":
    main()
